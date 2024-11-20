import re
import requests
from openpyxl.styles import PatternFill, Font
import time
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Border, Side

def getproducts(id):
    url = "https://market-backend.api.2gis.ru/5.0/product/items_by_branch"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 YaBrowser/24.10.0.0 Safari/537.36"
    }
    params = {
        "branch_id": f"{id}",
        "locale": "ru_RU",
        "page": "1",
        "page_size": "50",
        "feature_config": "categories_without_fake_first_level,range_price_type_supported,from_price_type_supported"
    }

    response = requests.get(url, headers=headers, params=params)

    # Проверка статуса и вывод результата
    if response.status_code == 200:
        data = response.json()

        return data


def getcontacts(id):
    result = {
        "phones": [],
        "emails": []
    }

    url = f"https://2gis.ru/firm/{id}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36"
    }

    start_time = time.time()

    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')

        # Поиск телефонов
        phone_elements = soup.find_all('a', class_='_2lcm958')
        for phone in phone_elements:
            href = phone.get('href')
            if href and href.startswith('tel:'):
                phone_number = href.replace('tel:', '')
                result['phones'].append(phone_number)

        # Поиск электронной почты
        email_elements = soup.find_all('a', class_='_2lcm958')
        for email in email_elements:
            href = email.get('href')
            if href and href.startswith('mailto:'):
                email_address = href.replace('mailto:', '')
                result['emails'].append(email_address)

    else:
        print(f"Не удалось получить доступ к странице. Статус-код: {response.status_code}")

    elapsed_time = time.time() - start_time
    print("Время выполнения запроса: {:.2f} секунд(ы)".format(elapsed_time))

    return result


def getregionsids(text, key):
    url = "https://catalog.api.2gis.com/2.0/region/search"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 YaBrowser/24.10.0.0 Safari/537.36"
    }
    params = {
        "q": f"{text}",
        "key": f"{key}",
    }

    response = requests.get(url, headers=headers, params=params)
    print(response.json())
    return response.json()


def get_parent_categories(region_id, key):
    url = "https://catalog.api.2gis.com/2.0/catalog/rubric/list"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 YaBrowser/24.10.0.0 Safari/537.36"
    }

    params = {
        "region_id": f"{region_id}",
        "key": f"{key}",
    }

    response = requests.get(url, headers=headers, params=params)

    return response.json()


def get_categories_by_parent_id(region_id, parent_id, key):
    url = "https://catalog.api.2gis.com/2.0/catalog/rubric/list"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 YaBrowser/24.10.0.0 Safari/537.36"
    }

    params = {
        "region_id": f"{region_id}",
        "parent_id": f"{parent_id}",
        "key": f"{key}",
    }

    response = requests.get(url, headers=headers, params=params)

    return response.json()


def get_org_by_rubric_id(rubric_id, region_id, page, key):
    url = "https://catalog.api.2gis.com/3.0/items"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 YaBrowser/24.10.0.0 Safari/537.36"
    }

    params = {
        "rubric_id": f"{rubric_id}",
        "region_id": f"{region_id}",
        "radius": "2000",
        "page": f"{page}",
        "key": f"{key}",
    }

    response = requests.get(url, headers=headers, params=params)

    return response.json()


def get_column_letter(col_num):
    """Превращает номер столбца в букву, поддерживает более 26 столбцов."""
    letters = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        letters = chr(remainder + 65) + letters
    return letters


def createXLSX(organization_products, organizations, selected_categories):
    workbook = openpyxl.Workbook()

    # Создаем стили для выравнивания и рамки
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Страница "ОБЩЕЕ"
    sheet_overview = workbook.active
    sheet_overview.title = "ОБЩЕЕ"

    # Заполнение заголовков
    sheet_overview["A1"] = "КАТЕГОРИИ"
    sheet_overview["B1"] = "КОЛ-ВО ОРГАНИЗАЦИЙ"
    sheet_overview["C1"] = "ВСЕГО КАТЕГОРИЙ"
    sheet_overview["D1"] = "ВСЕГО ОРГАНИЗАЦИЙ"

    for cell in sheet_overview["1:1"]:
        cell.alignment = center_alignment
        cell.border = thin_border

    # Заполнение данных
    total_categories = len(selected_categories)
    total_organizations = len(organizations)

    row = 2
    for category in selected_categories:
        sheet_overview[f"A{row}"] = category["name"]
        sheet_overview[f"B{row}"] = category["organizations"]
        sheet_overview[f"B{row}"].alignment = center_alignment
        sheet_overview[f"B{row}"].border = thin_border
        row += 1

    # Итоги
    sheet_overview[f"C2"] = total_categories
    sheet_overview[f"D2"] = total_organizations

    for cell in [sheet_overview["C2"], sheet_overview["D2"]]:
        cell.alignment = center_alignment
        cell.border = thin_border

    # Автоподбор ширины столбцов
    for col in sheet_overview.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        sheet_overview.column_dimensions[column].width = adjusted_width

    # Страница "Данные"
    sheet_data = workbook.create_sheet(title="Данные")

    headers = [
        "Категория", "Название организации", "Адрес", "Почта", "Телефон", "Ссылка", "Продукты"
    ]

    # Стили для заголовков
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Жёлтый фон
    header_font = Font(bold=True, color="000000")  # Жирный шрифт, чёрный цвет
    header_alignment = Alignment(horizontal="center", vertical="center")  # Выравнивание по центру
    header_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                           top=Side(style="thin"), bottom=Side(style="thin"))  # Рамка вокруг

    # Записываем заголовки в первую строку с применением стилей
    for col_num, header in enumerate(headers, 1):
        cell = sheet_data.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill  # Применяем цвет фона
        cell.font = header_font  # Применяем шрифт
        cell.alignment = header_alignment  # Применяем выравнивание
        cell.border = header_border  # Применяем рамку

    row = 2  # Начало с 2-й строки

    for org_id, org_data in organizations.items():
        category_name = next((cat["name"] for cat in selected_categories if cat["id"] == org_data["category_id"]),
                             "Неизвестно")

        address = org_data["address_name"] + " " + org_data.get("address_comment", "")
        email_phone = "; ".join(org_data['emails']) if org_data['emails'] else "-"
        phones = "; ".join(org_data['phones']) if org_data['phones'] else "-"

        products = organization_products.get(org_id, [])
        products_data = "; ".join([f"{product['name']} ({product['price']})" for product in products])

        org_link = f"https://2gis.ru/firm/{org_id}"

        sheet_data.cell(row=row, column=1, value=category_name)
        sheet_data.cell(row=row, column=2, value=org_data["name"])
        sheet_data.cell(row=row, column=3, value=address)
        sheet_data.cell(row=row, column=4, value=email_phone)
        sheet_data.cell(row=row, column=5, value=phones)
        sheet_data.cell(row=row, column=6, value=org_link)
        sheet_data.cell(row=row, column=7, value=products_data)

        row += 1

    # Автоподбор ширины столбцов
    for col in sheet_data.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        sheet_data.column_dimensions[column].width = adjusted_width

    workbook.save("organizations.xlsx")
